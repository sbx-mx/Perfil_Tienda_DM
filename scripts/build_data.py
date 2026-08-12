#!/usr/bin/env python3
"""Audita los motores y genera un contrato compacto para el dashboard."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import tempfile
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

MONTHS = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}
MONTH_LABELS = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
MAX_XLSX_UNCOMPRESSED = 180 * 1024 * 1024


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    ascii_text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^A-Z0-9]+", " ", ascii_text.upper()).strip()


def clean_cc(value: Any) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"): text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits[-5:].zfill(5) if digits else ""


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def as_iso(value: Any) -> str | None:
    if isinstance(value, (date, datetime)): return value.isoformat()[:10]
    text = str(value or "").strip()
    if not text or normalize(text) in {"NA", "N A", "N/A"}: return None
    for pattern in (r"^(\d{1,2})/(\d{1,2})/(\d{4})$", r"^(\d{4})-(\d{2})-(\d{2})"):
        match = re.match(pattern, text)
        if match:
            if pattern.startswith("^(\\d{4}"):
                return f"{match[1]}-{match[2]}-{match[3]}"
            return f"{match[3]}-{int(match[2]):02d}-{int(match[1]):02d}"
    return None


def number(value: Any, *, percent: bool = False) -> float | None:
    if value is None or isinstance(value, bool): return None
    if isinstance(value, (int, float)):
        result = float(value)
    else:
        text = str(value).strip()
        if not text or normalize(text) in {"NA", "N A", "N/A", "NULL"}: return None
        negative = text.startswith("(") and text.endswith(")")
        text = text.strip("()").replace("$", "").replace(",", "").replace("%", "")
        try: result = float(text)
        except ValueError: return None
        if negative: result = -result
        if percent: result /= 100
    return result if math.isfinite(result) else None


def round_number(value: float | None, digits: int = 6) -> float | None:
    if value is None: return None
    return round(value, digits)


def validate_xlsx(path: Path) -> None:
    if not path.is_file() or path.suffix.casefold() != ".xlsx" or not zipfile.is_zipfile(path):
        raise ValueError(f"XLSX inválido: {path.name}")
    with zipfile.ZipFile(path) as archive:
        entries = archive.infolist()
        total = sum(entry.file_size for entry in entries)
        if total > MAX_XLSX_UNCOMPRESSED: raise ValueError(f"XLSX excede límite seguro: {path.name}")
        for entry in entries:
            if entry.filename.startswith(("/", "\\")) or ".." in Path(entry.filename).parts:
                raise ValueError(f"Ruta interna insegura en {path.name}")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""): digest.update(chunk)
    return digest.hexdigest()


def atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as output:
            json.dump(payload, output, ensure_ascii=False, separators=(",", ":"), allow_nan=False)
            output.flush(); os.fsync(output.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def load_csv(path: Path, encoding: str = "utf-8-sig") -> list[dict[str, str]]:
    with path.open(encoding=encoding, newline="") as source:
        return list(csv.DictReader(source))


def load_export_csv(path: Path) -> list[dict[str, str]]:
    lines = path.read_text(encoding="utf-16").splitlines()
    try: start = next(index for index, line in enumerate(lines) if line.startswith('"Mes","'))
    except StopIteration as error: raise ValueError(f"No se encontró encabezado real en {path.name}") from error
    return list(csv.DictReader(lines[start:]))


def month_from_profile(value: Any) -> int | None:
    text = str(value or "").strip()
    if re.fullmatch(r"\d+(?:\.0)?", text):
        numeric = int(float(text))
        month = numeric % 100 if numeric >= 100000 else numeric
        return month if 1 <= month <= 12 else None
    match = re.match(r"^(?:\d+_)?([A-Za-zÁÉÍÓÚáéíóú]+)$", text)
    return MONTHS.get(normalize(match[1]).lower()) if match else None


def month_from_period(value: Any) -> int | None:
    text = str(value or "").strip()
    if not re.fullmatch(r"20\d{4}", text): return None
    month = int(text[-2:])
    return month if 1 <= month <= 12 else None


def year_from_period(value: Any) -> int | None:
    text = str(value or "").strip()
    return int(text[:4]) if re.fullmatch(r"20\d{4}", text) else None


def metric_format(header: str) -> str:
    key = normalize(header)
    if key in {"IPLH", "IPLH AA", "TPLH", "TPLH AA"}: return "decimal"
    if key in {"ICA SCORE", "ICA SCORE AA", "SEGUNDAS CONEXIONES", "SEGUNDAS CONEXIONES AA"}: return "number"
    if key in {"DT TIME", "TIEMPO DT AA"}: return "duration"
    return "percent"


def metric_direction(graph: str) -> str:
    key = normalize(graph)
    return "lower" if key in {"LABOR", "COSTO", "COSTO %", "DT TIME", "ROTACION"} else "higher"


def pair_kind(reference: str) -> str:
    return "ppto" if "PPTO" in normalize(reference) else "aa"


def build(root: Path, output: Path, audit_output: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    engines = root / "data" / "engines"
    paths = {
        "directory": engines / "Directorio_Perfil Tienda.csv",
        "profile": engines / "Base_Perfil Tienda.xlsx",
        "business_aa": engines / "Base_Perfil Tienda.csv",
        "business_real": engines / "Base_Perfil Tienda_2.csv",
        "mix_manifest": engines / "mix" / "manifest.json",
        "partners": engines / "Query.xlsx",
    }
    for path in paths.values():
        if not path.is_file(): raise ValueError(f"Falta motor: {path.name}")
    validate_xlsx(paths["profile"]); validate_xlsx(paths["partners"])

    issues: list[str] = []
    warnings: list[str] = []
    sources = {name: {"file": str(path.relative_to(engines)), "bytes": path.stat().st_size, "sha256": sha256(path)} for name, path in paths.items()}
    with paths["mix_manifest"].open(encoding="utf-8") as source:
        mix_manifest = json.load(source)
    mix_paths = [paths["mix_manifest"].parent / part["file"] for part in mix_manifest.get("parts", [])]
    if not mix_paths or any(not path.is_file() for path in mix_paths):
        raise ValueError("El manifiesto de Base_Mix no contiene todas sus partes")
    for path, part in zip(mix_paths, mix_manifest["parts"]):
        if path.stat().st_size != part["bytes"] or sha256(path) != part["sha256"]:
            raise ValueError(f"Parte Mix alterada o incompleta: {path.name}")
        sources[f"mix_{part['month']:02d}"] = {"file": str(path.relative_to(engines)), "bytes": path.stat().st_size, "sha256": part["sha256"]}

    directory_rows = load_csv(paths["directory"])
    directory: list[dict[str, Any]] = []
    directory_by_cc: dict[str, dict[str, Any]] = {}
    aliases: defaultdict[str, set[str]] = defaultdict(set)
    for index, row in enumerate(directory_rows, start=2):
        cc = clean_cc(row.get("CC"))
        if not re.fullmatch(r"\d{5}", cc):
            issues.append(f"Directorio fila {index}: CC inválido")
            continue
        if cc in directory_by_cc:
            issues.append(f"Directorio: CC duplicado {cc}")
            continue
        item = {
            "cc": cc,
            "store": clean_header(row.get("Tienda")),
            "region": clean_header(row.get("Región")),
            "dm": clean_header(row.get("DM")),
            "rd": clean_header(row.get("Director Regional")),
            "division": clean_header(row.get("Division")),
            "opened": as_iso(row.get("Fecha de Apertura")),
            "city": clean_header(row.get("Ciudad")),
            "state": clean_header(row.get("Estado")),
            "address": clean_header(row.get("Dirección")),
            "format": clean_header(row.get("Formato de Tienda\n (Tipo tienda 1)")),
            "generator": clean_header(row.get("Generador \n(Tipo tienda 2)")),
            "family": clean_header(row.get("Familia \n(Tipo tienda 3)")),
            "type5": clean_header(row.get("Tipo tienda 5")),
            "design": clean_header(row.get("Intensión del Diseño ")),
            "size": clean_header(row.get("Tamaño ")),
            "tier": clean_header(row.get("TIER")),
            "seats": number(row.get("# Seats in Store")),
            "manager": clean_header(row.get("Gerente")),
            "storeEmail": clean_header(row.get("Mail de Tienda")),
            "managerEmail": clean_header(row.get("Mail Gerente ")),
        }
        directory.append(item); directory_by_cc[cc] = item
        for alias in (row.get("Tienda"), row.get("Nombre APP y Signage")):
            if normalize(alias): aliases[normalize(alias)].add(cc)

    workbook = load_workbook(paths["profile"], read_only=True, data_only=True, keep_links=False)
    if "Perfil" not in workbook.sheetnames or "Instrucciones_Ejemplo" not in workbook.sheetnames:
        raise ValueError("Base_Perfil Tienda.xlsx requiere Perfil e Instrucciones_Ejemplo")
    profile_sheet = workbook["Perfil"]
    headers = [clean_header(value) for value in next(profile_sheet.iter_rows(values_only=True))]
    normalized_headers = [normalize(header) for header in headers]
    month_aliases = {"MES", "MES NUM", "MES NUMERO", "PERIODO"}
    try: month_column = next(index for index, header in enumerate(normalized_headers) if header in month_aliases)
    except StopIteration as error: raise ValueError("Perfil requiere MES_NUM, Mes o Periodo") from error
    try: cc_column = next(index for index, header in enumerate(normalized_headers) if header in {"CECO", "CC", "TIENDA", "TIENDAS"})
    except StopIteration as error: raise ValueError("Perfil requiere CeCo") from error
    if month_column == cc_column: raise ValueError("Mes y CeCo no pueden compartir columna")
    duplicate_headers = [header for header, count in Counter(headers).items() if count > 1]
    if duplicate_headers: raise ValueError(f"Encabezados duplicados en Perfil: {duplicate_headers}")

    instruction_rows = []
    for row in workbook["Instrucciones_Ejemplo"].iter_rows(min_row=2, values_only=True):
        values = [clean_header(value) for value in row[:5]]
        if any(values): instruction_rows.append({"pillar": values[0], "header": values[1], "graph": values[2], "instruction": values[3], "note": values[4]})
    instruction_by_header = {normalize(item["header"]): item for item in instruction_rows if item["header"]}

    metric_columns = [index for index in range(len(headers)) if index not in {month_column, cc_column}]
    metric_headers = [headers[index] for index in metric_columns]
    metric_index = {header: index for index, header in enumerate(metric_headers)}
    graphs: list[dict[str, Any]] = []
    seen_graphs: set[tuple[str, str]] = set()
    visible_instructions = [item for item in instruction_rows if normalize(item["graph"]) not in {"", "NO SE MUESTRA"} and normalize(item["pillar"]) in {"PARTNER", "CLIENTE", "NEGOCIO"}]
    header_aliases = {"ROTACION": "ROLLING RY", "ROTACION AA": "ROLLING RY AA"}
    for item in visible_instructions:
        graph_key = (normalize(item["pillar"]), normalize(item["graph"]))
        if graph_key in seen_graphs: continue
        seen_graphs.add(graph_key)
        candidates = [entry["header"] for entry in visible_instructions if (normalize(entry["pillar"]), normalize(entry["graph"])) == graph_key]
        actual = next((header for header in candidates if " AA" not in f" {normalize(header)}" and "PPTO" not in normalize(header)), candidates[0])
        reference = next((header for header in candidates if normalize(header) != normalize(actual)), "")
        actual_key = header_aliases.get(normalize(actual), normalize(actual))
        reference_key = header_aliases.get(normalize(reference), normalize(reference))
        actual = next((header for header in metric_headers if normalize(header) == actual_key), actual)
        reference = next((header for header in metric_headers if normalize(header) == reference_key), reference)
        if actual not in metric_index:
            warnings.append(f"Métrica clasificada no encontrada: {actual}")
            continue
        graph_name = normalize(item["graph"])
        metric_name = normalize(actual)
        # VMT y OMT son indicadores operativos de Negocio aunque la hoja de
        # clasificación de versiones anteriores los ubicara en Cliente.
        pillar = "Negocio" if graph_name in {"VMT", "OMT"} or metric_name in {"VMT", "OMT"} else item["pillar"]
        graphs.append({
            "id": re.sub(r"[^a-z0-9]+", "-", normalize(item["graph"]).lower()).strip("-"),
            "pillar": pillar, "title": item["graph"], "actual": actual,
            "reference": reference if reference in metric_index else None,
            "referenceKind": pair_kind(reference), "format": metric_format(actual),
            "direction": metric_direction(item["graph"]),
            "ytd": "latest" if normalize(item["graph"]) == "ROTACION" else "average",
            "instruction": item["instruction"], "note": item["note"],
        })

    profile: defaultdict[str, dict[str, list[float | None]]] = defaultdict(dict)
    profile_seen: set[tuple[str, int]] = set()
    profile_unmatched = Counter(); minus_100_blanked = Counter(); duration_blanked = Counter(); profile_months = Counter()
    for row_number, row in enumerate(profile_sheet.iter_rows(min_row=2, values_only=True), start=2):
        month = month_from_profile(row[month_column]); cc = clean_cc(row[cc_column])
        if month is None or not cc:
            warnings.append(f"Perfil fila {row_number}: periodo o CeCo inválido")
            continue
        if cc not in directory_by_cc:
            profile_unmatched[cc] += 1; continue
        key = (cc, month)
        if key in profile_seen:
            issues.append(f"Perfil duplicado: {cc}, mes {month}"); continue
        profile_seen.add(key); profile_months[month] += 1
        values: list[float | None] = []
        for header, column in zip(metric_headers, metric_columns):
            raw = row[column] if column < len(row) else None
            value = number(raw)
            if metric_format(header) == "percent" and value is not None and math.isclose(value, -1.0, abs_tol=1e-9):
                minus_100_blanked[header] += 1; value = None
            if metric_format(header) == "duration" and value is not None and abs(value) < 1:
                # El motor exporta DT como fracción de día que Excel interpreta
                # visualmente como hh:mm. El negocio lo define como mm:ss:
                # 0.052083 (01:15 en origen) debe representar 75 segundos.
                value *= 1440
            if metric_format(header) == "duration" and value is not None and not 20 <= value <= 1800:
                duration_blanked[header] += 1; value = None
            values.append(round_number(value))
        profile[cc][str(month)] = values

    business: defaultdict[str, dict[str, dict[str, float | None]]] = defaultdict(dict)
    business_unmatched = Counter(); business_duplicates = Counter(); business_months = Counter(); business_years = Counter()
    business_inputs = {name: load_export_csv(paths[name]) for name in ("business_aa", "business_real")}
    business_real_headers = list(business_inputs["business_real"][0]) if business_inputs["business_real"] else []
    required_business = {
        "business_aa": {"Mes", "Tiendas", "ADT AA", "OMT"},
        "business_real": {"Mes", "Tiendas", "ADT Real", "Venta $", "Var Ventas vs Ppto (%)", "AWS $", "Ticket Prom Real", "Ticket Prom AA", "Ticket Prom Ppto", "Var Ticket vs AA (%)", "Var Ticket vs Ppto (%)"},
    }
    for source_name, rows in business_inputs.items():
        source_headers = set(rows[0]) if rows else set()
        if missing := required_business[source_name].difference(source_headers):
            raise ValueError(f"{paths[source_name].name} sin encabezados requeridos: {sorted(missing)}")
    business_seen: set[tuple[str, int, str]] = set()
    for source_name, rows in business_inputs.items():
        for row in rows:
            month = month_from_period(row.get("Mes")); year = year_from_period(row.get("Mes")); cc = clean_cc(row.get("Tiendas"))
            if month is None or not cc: continue
            if year is not None: business_years[year] += 1
            if cc not in directory_by_cc:
                business_unmatched[cc] += 1; continue
            target = business[cc].setdefault(str(month), {})
            source_key = (cc, month, source_name)
            if source_key in business_seen:
                business_duplicates[f"{cc}-{month}-{source_name}"] += 1
                continue
            business_seen.add(source_key)
            if source_name == "business_aa":
                target.update({"adtAa": round_number(number(row.get("ADT AA"))), "omtDiff": round_number(number(row.get("OMT")))})
            else:
                sales = number(row.get("Venta $")); variance = number(row.get("Var Ventas vs Ppto (%)"), percent=True)
                budget = sales / (1 + variance) if sales is not None and variance is not None and not math.isclose(variance, -1) else None
                target.update({
                    "adt": round_number(number(row.get("ADT Real"))), "sales": round_number(sales, 2),
                    "salesBudget": round_number(budget, 2), "salesVariance": round_number(variance),
                    "aws": round_number(number(row.get("AWS $")), 2), "ticket": round_number(number(row.get("Ticket Prom Real")), 2),
                    "ticketAa": round_number(number(row.get("Ticket Prom AA")), 2),
                    "ticketBudget": round_number(number(row.get("Ticket Prom Ppto")), 2),
                    "ticketVariance": round_number(number(row.get("Var Ticket vs AA (%)"), percent=True)),
                    "ticketBudgetVariance": round_number(number(row.get("Var Ticket vs Ppto (%)"), percent=True)),
                })
            business_months[month] += 1

    alias_to_cc = {alias: next(iter(ccs)) for alias, ccs in aliases.items() if len(ccs) == 1}
    mix: defaultdict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    mix_unmatched = Counter(); mix_invalid_sales = 0; mix_rows = 0; mix_matched_rows = 0
    mix_months = Counter()
    for mix_path in mix_paths:
        with mix_path.open(encoding="utf-8-sig", newline="") as source:
            for row in csv.DictReader(source):
                mix_rows += 1; month = month_from_profile(row.get("Mes"))
                cc = alias_to_cc.get(normalize(row.get("Tienda")))
                if not cc:
                    mix_unmatched[clean_header(row.get("Tienda"))] += 1; continue
                sale = number(row.get("Venta"))
                if month is None or sale is None:
                    mix_invalid_sales += 1; continue
                mix_matched_rows += 1; mix_months[month] += 1
                target = mix[cc].setdefault(str(month), {"category": defaultdict(float), "order": defaultdict(float), "total": 0.0})
                category = clean_header(row.get("Category")) or "Sin categoría"
                order = clean_header(row.get("Tipo Orden")) or "Sin canal"
                target["category"][category] += sale; target["order"][order] += sale; target["total"] += sale
    for months in mix.values():
        for target in months.values():
            total = target["total"]
            target["category"] = {key: round_number(value / total) for key, value in target["category"].items()} if total else {}
            target["order"] = {key: round_number(value / total) for key, value in target["order"].items()} if total else {}
            target["total"] = round_number(total, 4)

    partner_book = load_workbook(paths["partners"], read_only=True, data_only=True, keep_links=False)
    if "Query" not in partner_book.sheetnames: raise ValueError("Query.xlsx requiere pestaña Query")
    if "Instrucciones" in partner_book.sheetnames:
        for row in partner_book["Instrucciones"].iter_rows(values_only=True):
            header, instruction = (clean_header(value) for value in row[:2])
            if header or instruction:
                instruction_rows.append({"pillar": "Query Partner", "header": header, "graph": "Uso", "instruction": instruction, "note": ""})
    query = partner_book["Query"]
    query_headers = [clean_header(value) for value in next(query.iter_rows(values_only=True))]
    required_query = {"NUM_EMP", "NOM_PUESTO", "SEXO", "F.NAC", "F_INGRESO", "cc", "STATUS_ EMP (ACTIVO/BAJA)"}
    if missing := required_query.difference(query_headers): raise ValueError(f"Query.xlsx sin encabezados: {sorted(missing)}")
    qi = {header: query_headers.index(header) for header in required_query}
    employees: dict[tuple[str, str], tuple[Any, ...]] = {}; partner_unmatched = Counter()
    for row in query.iter_rows(min_row=2, values_only=True):
        cc = clean_cc(row[qi["cc"]]); employee = str(row[qi["NUM_EMP"]] or "").strip()
        if not cc or not employee: continue
        if cc not in directory_by_cc: partner_unmatched[cc] += 1; continue
        employees[(cc, employee)] = row
    partner_groups: defaultdict[str, list[tuple[Any, ...]]] = defaultdict(list)
    for (cc, _), row in employees.items(): partner_groups[cc].append(row)
    partners: dict[str, dict[str, Any]] = {}
    today = datetime.now(UTC).date()
    for cc, rows in partner_groups.items():
        active = [row for row in rows if "ACTIV" in normalize(row[qi["STATUS_ EMP (ACTIVO/BAJA)"]])]
        roles = Counter(clean_header(row[qi["NOM_PUESTO"]]) or "Sin puesto" for row in active)
        sex = Counter(normalize(row[qi["SEXO"]]) for row in active)
        ages = []; tenures = []; birthday_month = 0; anniversary_month = 0
        for row in active:
            birth = row[qi["F.NAC"]]; joined = row[qi["F_INGRESO"]]
            if isinstance(birth, (date, datetime)):
                birth_date = birth.date() if isinstance(birth, datetime) else birth
                ages.append(today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day)))
                birthday_month += int(birth_date.month == today.month)
            if isinstance(joined, (date, datetime)):
                joined_date = joined.date() if isinstance(joined, datetime) else joined
                tenures.append(max(0, (today.year - joined_date.year) * 12 + today.month - joined_date.month))
                anniversary_month += int(joined_date.month == today.month)
        partners[cc] = {
            "headcount": len(active), "baristas": sum(value for key, value in roles.items() if "BARISTA" in normalize(key)),
            "supervisors": sum(value for key, value in roles.items() if "SUPERVISOR" in normalize(key)),
            "managers": sum(value for key, value in roles.items() if any(term in normalize(key) for term in ("GERENTE", "STORE MANAGER", "SUBGERENTE"))),
            "female": sex.get("F", 0), "male": sex.get("M", 0),
            "avgAge": round_number(sum(ages) / len(ages), 1) if ages else None,
            "avgTenureMonths": round_number(sum(tenures) / len(tenures), 1) if tenures else None,
            "birthdaysThisMonth": birthday_month, "anniversariesThisMonth": anniversary_month,
            "roles": dict(roles.most_common(8)),
        }

    all_months = sorted({int(month) for values in profile.values() for month in values} | {int(month) for values in business.values() for month in values} | {int(month) for values in mix.values() for month in values})
    coverage = []
    for item in directory:
        cc = item["cc"]
        coverage.append({"cc": cc, "profile": cc in profile, "business": cc in business, "mix": cc in mix, "partners": cc in partners})
    for count, message in (
        (len(profile_unmatched), "Perfil: {count} CeCo sin cruce; se dejaron en blanco."),
        (len(business_unmatched), "Negocio: {count} CeCo sin cruce; se dejaron en blanco."),
        (len(mix_unmatched), "Mix: {count} nombres sin coincidencia exacta única; se dejaron en blanco."),
        (len(partner_unmatched), "Query: {count} CeCo sin cruce; se dejaron en blanco."),
    ):
        if count: warnings.append(message.format(count=count))
    if business_duplicates: issues.append(f"Negocio contiene {sum(business_duplicates.values())} claves duplicadas")

    audit = {
        "schemaVersion": 2, "generatedAt": datetime.now(UTC).isoformat(), "issueCount": len(issues), "warningCount": len(warnings),
        "issues": issues, "warnings": warnings, "sources": sources,
        "directory": {"rows": len(directory_rows), "validStores": len(directory)},
        "profile": {"rows": profile_sheet.max_row - 1, "matchedStores": len(profile), "months": dict(profile_months), "monthHeader": headers[month_column], "ccHeader": headers[cc_column], "minus100Blanked": dict(minus_100_blanked), "durationOutliersBlanked": dict(duration_blanked), "unmatched": dict(profile_unmatched)},
        "business": {"matchedStores": len(business), "months": dict(business_months), "years": dict(business_years), "unmatched": dict(business_unmatched), "realHeaders": business_real_headers},
        "mix": {"sourceRows": mix_rows, "manifestRows": mix_manifest.get("rows"), "parts": len(mix_paths), "months": dict(mix_months), "matchedRows": mix_matched_rows, "matchedStores": len(mix), "invalidRows": mix_invalid_sales, "unmatchedNames": len(mix_unmatched), "unmatchedTop": dict(mix_unmatched.most_common(100))},
        "partners": {"uniqueEmployees": len(employees), "matchedStores": len(partners), "unmatched": dict(partner_unmatched)},
        "coverage": coverage,
    }
    if issues: raise ValueError(" | ".join(issues))

    display_year = max(business_years) if business_years else datetime.now(UTC).year
    payload = {
        "schemaVersion": 2, "generatedAt": audit["generatedAt"],
        "months": [{"id": month, "period": f"{display_year}{month:02d}", "label": MONTH_LABELS[month - 1], "short": MONTH_LABELS[month - 1][:3]} for month in all_months],
        "directory": directory, "metricHeaders": metric_headers, "graphs": graphs,
        "profile": profile, "business": business, "mix": mix, "partners": partners,
        "instructions": instruction_rows, "auditSummary": {"warnings": len(warnings), "generatedAt": audit["generatedAt"]},
    }
    atomic_json(output, payload); atomic_json(audit_output, audit)
    return payload, audit


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--output", type=Path)
    parser.add_argument("--audit-output", type=Path)
    args = parser.parse_args(); root = args.root.resolve()
    output = args.output.resolve() if args.output else root / "data" / "dashboard.json"
    audit_output = args.audit_output.resolve() if args.audit_output else root / "data" / "audit.json"
    try: payload, audit = build(root, output, audit_output)
    except (OSError, ValueError, zipfile.BadZipFile) as error: raise SystemExit(f"Construcción cancelada: {error}") from error
    print(json.dumps({"status": "ready", "stores": len(payload["directory"]), "months": len(payload["months"]), "warnings": audit["warningCount"], "output": str(output)}, ensure_ascii=False))


if __name__ == "__main__": main()
